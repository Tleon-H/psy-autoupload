from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.select import Select
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import time
import pandas as pd
import glob
import os
import sys
from datetime import datetime

def resource_path(relative_path):
    """获取资源文件路径"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class FormExtractor:
    def __init__(self):
        self.driver = None
        self.login_url = 'https://i.wzut.edu.cn/_s2/teachers_sy/main.psp'
        self.is_logged_in = False
        self.form_data = None
        self.student_list = None
        self.excel_file = None  # 添加 excel_file 属性
        self.data = None  # 添加 data 属性
        # 尝试从配置文件加载数据
        self._load_config()
        
        # 检查并创建必要的目录和文件
        self.check_environment()
        
    def _load_config(self):
        """从配置文件加载数据"""
        try:
            with open('form_config.json', 'r', encoding='utf-8') as f:
                config = json.load(f)
                if 'fields' in config:
                    self.form_data = config
                if 'student_list' in config:
                    self.student_list = config['student_list']
        except:
            pass
    
    def login(self):
        """登录系统"""
        try:
            if not self.driver:
                print('正在启动浏览器...')
                # 添加 Chrome 选项
                options = Options()
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument('--allow-insecure-localhost')
                options.add_argument('--allow-running-insecure-content')
                # 禁用 SSL 检查
                options.add_experimental_option('excludeSwitches', ['enable-logging'])
                options.set_capability('acceptInsecureCerts', True)
                
                # 使用选项创建 driver
                self.driver = webdriver.Chrome(
                    service=Service(ChromeDriverManager().install()),
                    options=options
                )
                # 设置页面加载超时
                self.driver.set_page_load_timeout(30)
            
            if not self.is_logged_in:
                print('正在打开登录页面...')
                self.driver.get(self.login_url)
                
                print('- 开始自动登录...')
                # 等待登录页面加载
                username_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, "username"))
                )
                password_input = self.driver.find_element(By.ID, "password")
                
                # 输入账号密码
                username_input.send_keys("20210254")
                password_input.send_keys("Htl887280")
                time.sleep(1)
                
                print('\n请手动完成登录验证，完成后按回车键继续...')
                input()
                
                # 等待登录成功
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//a[contains(text(), '学工系统')]"))
                )
                print('✓ 登录成功')
                self.is_logged_in = True
            
            return True
            
        except Exception as e:
            print(f'登录失败: {str(e)}')
            return False
    
    def return_to_home(self):
        """返回到登录后的主页面"""
        try:
            print('\n正在返回主页面...')
            # 关闭当前窗口
            if len(self.driver.window_handles) > 1:
                self.driver.close()
                # 切换回主窗口
                self.driver.switch_to.window(self.driver.window_handles[0])
            else:
                # 如果只有一个窗口，直接返回主页
                self.driver.get(self.login_url)
            time.sleep(1)
            
            # 验证是否仍然登录
            try:
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//a[contains(text(), '学工系统')]"))
                )
            except:
                print('登录已失效，重新登录...')
                self.is_logged_in = False
                return self.login()
                
            print('✓ 已返回主页面')
            return True
            
        except Exception as e:
            print(f'返回主页面失败: {str(e)}')
            return False
    
    def _navigate_to_form(self):
        """导航到谈心记录表单页面"""
        try:
            print('正在导航到表单页面...')
            
            # 1. 点击学工系统
            print('- 点击学工系统...')
            xgxt_link = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '学工系统')]"))
            )
            xgxt_link.click()
            time.sleep(2)
            
            # 切换到新窗口
            print('- 切换到学工系统窗口...')
            self.driver.switch_to.window(self.driver.window_handles[-1])
            
            # 2. 点击学工管理系统
            print('- 点击学工管理系统...')
            xggl_link = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '学工管理系统')]"))
            )
            xggl_link.click()
            time.sleep(2)
            
            # 直接导航到表单页面
            print('- 直接导航到表单页面...')
            form_url = 'https://xgxt.wzut.edu.cn/wzutXG/Sys/SystemForm/FDYWork/FDYTalkHeartManageEdit.aspx?Status=Add&type=3'
            self.driver.get(form_url)
            time.sleep(2)
            
            print('✓ 已进入表单页面')
            return True
            
        except Exception as e:
            print(f'导航失败: {str(e)}')
            return False
    
    def extract_form_fields(self):
        """提取表单字段信息"""
        try:
            if not self._navigate_to_form():
                return False
            
            # 等待主表单加载
            print('正在查找表单...')
            form = WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.TAG_NAME, 'form'))
            )
            print('✓ 找到主表单')
            
            # 提取基本表单信息
            form_data = {
                'form_id': form.get_attribute('id'),
                'form_action': form.get_attribute('action'),
                'fields': {}
            }
            
            # 提取主表单字段
            _extract_fields(self.driver, form, form_data['fields'])
            
            # 处理动态表单
            print('\n检查动态表单...')
            try:
                # 等待谈话方式下拉框加载
                talk_type = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.NAME, 'TalkHeartPage$TalkType1'))
                )
                
                # 使用 Select 类来选择选项
                select = Select(talk_type)
                select.select_by_value('08031001')  # 一对一谈话的值
                time.sleep(3)
                
                print('已选择"一对一谈话"选项')
                
                # 等待动态字段加载
                time.sleep(2)
                
                # 重新获取所有字段
                all_fields = form.find_elements(By.CSS_SELECTOR, 
                    'input:not([type="hidden"]), select, textarea')
                
                # 提取所有字段信息
                _extract_fields(self.driver, form, form_data['fields'], all_fields)
                
            except Exception as e:
                print(f'提取动态字段时出错: {str(e)}')
                print('继续保存已提取的基本字段')
            
            # 读取现有配置以保留学生名单
            try:
                with open('form_config.json', 'r', encoding='utf-8') as f:
                    existing_config = json.load(f)
                    if 'student_list' in existing_config:
                        form_data['student_list'] = existing_config['student_list']
                        print('✓ 保留已有的学生名单')
            except:
                print('未找到现有配置文件，将创建新文件')
            
            # 保存结果
            with open('form_config.json', 'w', encoding='utf-8') as f:
                json.dump(form_data, f, ensure_ascii=False, indent=2)
            
            print(f'\n✓ 已提取 {len(form_data["fields"])} 个字段')
            if 'student_list' in form_data:
                print(f'✓ 保留 {len(form_data["student_list"])} 条学生记录')
            print('✓ 数据已保存到 form_config.json')
            
            # 保存到类成员变量
            self.form_data = form_data
            return True
            
        except Exception as e:
            print(f'提取表单字段失败: {str(e)}')
            return False
        finally:
            self.return_to_home()
    
    def _load_excel_data(self):
        """加载Excel数据"""
        try:
            print(f'正在加载数据: {self.excel_file}')
            df = pd.read_excel(self.excel_file)
            
            # 获取字段映射和记录数据
            field_map = dict(zip(df['字段名称'], df['系统字段']))
            record_cols = [col for col in df.columns if col.startswith('记录')]
            
            if not record_cols:
                print('❌ 未找到记录数据')
                return False
            
            # 转换数据格式
            records = []
            for col in record_cols:
                record = {}
                for _, row in df.iterrows():
                    if pd.notna(row[col]):
                        record[field_map[row['字段名称']]] = str(row[col])
                if record:
                    records.append(record)
            
            print(f'✓ 已加载 {len(records)} 条记录')
            self.data = records
            return True
            
        except Exception as e:
            print(f'加载Excel数据失败: {str(e)}')
            return False
    
    def _handle_student_name(self, name):
        """处理学生姓名输入和查询"""
        try:
            # 等待姓名输入框可见
            name_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, 'TalkHeartPage$StuName'))
            )
            name_input.clear()
            name_input.send_keys(name)
            time.sleep(1)  # 等待输入完成
            
            # 点击姓名右边的查询按钮
            print('  - 点击查询按钮...')
            search_btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, 'TalkHeartPage_btnSearchStuInfo1'))
            )
            search_btn.click()
            time.sleep(2)  # 等待查询结果
            
            # 验证查询结果
            student_id = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.ID, 'TalkHeartPage_StuId'))
            )
            if not student_id.get_attribute('value'):
                print('⚠️ 未找到学生信息，请检查姓名是否正确')
                return False
            print('  ✓ 学生信息查询成功')
            return True
            
        except Exception as e:
            print(f'处理学生姓名时出错: {str(e)}')
            return False

    def _handle_conversation_topics(self, value):
        """处理谈话主题多选"""
        try:
            # 将输入的多选值拆分为列表
            topics = [t.strip() for t in value.split(',')]
            
            # 遍历所有主题复选框
            for i in range(7):
                field_name = f'TalkHeartPage$ConversationTopic${i}'
                try:
                    checkbox = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.NAME, field_name))
                    )
                    # 获取当前复选框对应的标签文本
                    label = self.driver.find_element(
                        By.XPATH, 
                        f"//input[@name='{field_name}']/following-sibling::label"
                    ).text
                    
                    # 如果标签文本在选中列表中，点击复选框
                    if label in topics:
                        if not checkbox.is_selected():
                            checkbox.click()
                    else:
                        if checkbox.is_selected():
                            checkbox.click()
                except:
                    continue
                    
            # 验证是否至少选择了一项
            selected = False
            for i in range(7):
                try:
                    checkbox = self.driver.find_element(By.NAME, f'TalkHeartPage$ConversationTopic${i}')
                    if checkbox.is_selected():
                        selected = True
                        break
                except:
                    continue
                    
            if not selected:
                print('⚠️ 警告：未选择任何谈话主题')
                return False
                
            return True
            
        except Exception as e:
            print(f'处理谈话主题时出错: {str(e)}')
            return False

    def _handle_radio(self, field, value):
        """处理单选框"""
        radio_value = 'IsCareY' if value.lower() in ['是', 'true', '1', 'yes'] else 'IsCareN'
        radio = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, f'input[name="{field}"][value="{radio_value}"]'))
        )
        radio.click()

    def _fill_field(self, field, value):
        """填写普通字段"""
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.NAME, field))
        )
        
        # 特殊处理日期字段
        if field == 'TalkHeartPage$FinishDate':
            try:
                # 使用 JavaScript 设置日期值
                date_str = str(value)[:10]  # 只取日期部分 YYYY-MM-DD
                self.driver.execute_script(
                    "arguments[0].value = arguments[1];", 
                    element, 
                    date_str
                )
                # 触发 change 事件
                self.driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));",
                    element
                )
                return
            except Exception as e:
                print(f'设置日期时出错: {str(e)}')
                raise
        
        # 处理其他字段
        if element.tag_name == 'select':
            Select(element).select_by_visible_text(str(value))
        else:
            try:
                # 先尝试直接设置
                element.clear()
                element.send_keys(str(value))
            except:
                # 如果失败，尝试使用 JavaScript 设置值
                self.driver.execute_script(
                    "arguments[0].value = arguments[1];", 
                    element, 
                    str(value)
                )

    def _navigate_to_student_list(self):
        """导航到学生名单页面"""
        try:
            print('正在导航到名单页面...')
            
            # 1. 点击学工系统
            print('- 点击学工系统...')
            xgxt_link = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '学工系统')]"))
            )
            xgxt_link.click()
            time.sleep(1)
            
            # 切换到新窗口
            print('- 切换到学工系统窗口...')
            self.driver.switch_to.window(self.driver.window_handles[-1])
            
            # 直接导航到名单页面
            print('- 直接导航到名单页面...')
            list_url = 'https://xgxt.wzut.edu.cn/wzutXG/Sys/SystemForm/HeartHealth/FDYUploadList.aspx'
            self.driver.get(list_url)
            time.sleep(2)
            
            print('✓ 已进入名单页面')
            print('\n请在页面上选择要提取的学期，选择完成后按回车键继续...')
            input()
            return True
            
        except Exception as e:
            print(f'导航失败: {str(e)}')
            return False

    def extract_student_list(self):
        """提取学生名单"""
        try:
            if not self._navigate_to_student_list():
                return False
            
            # 等待表格加载
            print('正在提取学生名单...')
            try:
                # 先等待页面加载完成
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'form1'))
                )
                time.sleep(2)  # 等待表格数据加载
                
                # 尝试多种方式定位表格
                table = None
                try:
                    table = self.driver.find_element(By.ID, 'GridView1')
                except:
                    try:
                        table = self.driver.find_element(By.CSS_SELECTOR, '.GridViewStyle')
                    except:
                        try:
                            table = self.driver.find_element(By.XPATH, "//table[contains(@class, 'GridView')]")
                        except:
                            raise Exception('未找到学生名单表格，请确保已选择正确的学期')
                
                if not table:
                    raise Exception('未找到学生名单表格')
                
                # 提取表头
                headers = []
                header_cells = table.find_elements(By.TAG_NAME, 'th')
                if not header_cells:  # 如果没有th，尝试找第一行的td
                    header_cells = table.find_element(By.TAG_NAME, 'tr').find_elements(By.TAG_NAME, 'td')
                
                for cell in header_cells:
                    headers.append(cell.text.strip())
                
                if not headers:
                    raise Exception('未找到表格表头')
                    
                print(f'找到表头: {headers}')
                
                # 提取所有页面的学生信息
                student_list = []
                page = 1
                
                while True:
                    print(f'\n正在提取第 {page} 页...')
                    
                    # 提取当前页的数据
                    rows = table.find_elements(By.TAG_NAME, 'tr')[1:]  # 跳过表头行
                    current_page_count = 0
                    
                    for row in rows:
                        student_info = {}
                        cells = row.find_elements(By.TAG_NAME, 'td')
                        if len(cells) >= len(headers):  # 确保单元格数量正确
                            for i, cell in enumerate(cells):
                                if i < len(headers):
                                    value = cell.text.strip()
                                    if value:  # 只保存非空值
                                        student_info[headers[i]] = value
                            if student_info:  # 只添加非空记录
                                student_list.append(student_info)
                                current_page_count += 1
                    
                    print(f'✓ 本页提取了 {current_page_count} 条记录')
                    
                    # 尝试翻到下一页
                    try:
                        # 查找下一页按钮
                        next_button = self.driver.find_element(By.XPATH, "//a[text()='下一页']")
                        if 'disabled' in next_button.get_attribute('class'):
                            print('已到达最后一页')
                            break
                            
                        # 点击下一页
                        next_button.click()
                        time.sleep(2)  # 等待新页面加载
                        
                        # 重新获取表格元素（因为页面已刷新）
                        table = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.ID, 'GridView1'))
                        )
                        page += 1
                        
                    except Exception as e:
                        print('未找到下一页按钮或已到达最后一页')
                        break
                
                if not student_list:
                    raise Exception('未找到学生信息')
                    
                print(f'\n✓ 总共提取了 {len(student_list)} 名学生信息')
                
                # 打印第一条记录作为示例
                if student_list:
                    print('\n示例数据:')
                    for field, value in student_list[0].items():
                        print(f'{field}: {value}')
                
                # 读取现有配置
                try:
                    with open('form_config.json', 'r', encoding='utf-8') as f:
                        config = json.load(f)
                except:
                    config = {}
                
                # 更新学生名单
                config['student_list'] = student_list
                
                # 保存配置
                with open('form_config.json', 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                
                print('✓ 学生名单已保存到配置文件')
                # 保存到类成员变量
                self.student_list = student_list
                # 更新表单数据
                if config.get('fields'):
                    self.form_data = config
                return True
                
            except Exception as e:
                print(f'提取表格数据失败: {str(e)}')
                # 保存页面源码以便调试
                with open('page_source.html', 'w', encoding='utf-8') as f:
                    f.write(self.driver.page_source)
                print('页面源码已保存到 page_source.html')
                return False
            
        except Exception as e:
            print(f'提取学生名单失败: {str(e)}')
            return False
        finally:
            self.return_to_home()

    def generate_template_by_care_level(self):
        """根据关注等级生成不同的Excel模板"""
        try:
            print('正在生成谈心记录模板...')
            
            # 检查是否已提取表单字段
            if not self.form_data:
                print('❌ 未找到表单字段信息')
                print('请先提取表单字段')
                return False
            
            # 检查是否已提取学生名单
            if not self.student_list:
                print('❌ 未找到学生名单')
                print('请先提取学生名单')
                return False
            
            print(f'找到 {len(self.student_list)} 名学生')
            
            # 按关注等级分组
            care_levels = {
                '重点（二级）': [],
                '日常（三级）': [],
                '一般（四级）': []
            }
            
            # 分类学生
            for student in self.student_list:
                # 从学生信息中提取关注等级
                care_level = None
                for key in student:
                    if '关注等级' in key:
                        care_level = student[key]
                        break
            
                # 将学生添加到对应分组
                if care_level in care_levels:
                    care_levels[care_level].append(student)
                else:
                    print(f'未知关注等级: {care_level}，学生: {student.get("姓名", "未知")}')
            
            # 打印分类统计
            print('\n学生分类统计：')
            for level, students in care_levels.items():
                print(f'- {level}: {len(students)} 人')
            
            # 为每个关注等级生成模板
            generated_count = 0
            for level, students in care_levels.items():
                if students:  # 只为有学生的等级生成文件
                    output_file = f'谈心记录模板_{level}.xlsx'
                    if self._generate_template(output_file, len(students), students):
                        print(f'✓ 已生成{level}学生模板: {output_file}')
                        print(f'  - 包含 {len(students)} 名学生')
                        generated_count += 1
            
            if generated_count > 0:
                print(f'\n✓ 成功生成 {generated_count} 个模板文件')
                return True
            else:
                print('\n❌ 未能生成任何模板文件')
                return False
            
        except Exception as e:
            print(f'生成模板失败: {str(e)}')
            return False
    
    def _generate_template(self, output_file, record_count, student_list):
        """生成谈心记录模板"""
        try:
            # 准备数据
            template_data = []
            
            # 基本信息（必填）
            template_data.extend([
                {
                    '字段名称': '谈话日期',
                    '字段类型': '日期',
                    '必填': '是',
                    '可选值': '',
                    '填写说明': '格式：YYYY-MM-DD',
                    '系统字段': 'TalkHeartPage$FinishDate'
                },
                {
                    '字段名称': '学期',
                    '字段类型': '下拉选择',
                    '必填': '是',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$TermNo']['options'] if opt['text']]),
                    '填写说明': '选择当前学期',
                    '系统字段': 'TalkHeartPage$TermNo'
                }
            ])
            
            # 学生信息（必填）
            template_data.extend([
                {
                    '字段名称': '姓名',
                    '字段类型': '文本',
                    '必填': '是',
                    '可选值': '',
                    '填写说明': '输入学生姓名后会自动查询',
                    '系统字段': 'TalkHeartPage$StuName'
                }
            ])
            
            # 谈话信息（必填）
            template_data.extend([
                {
                    '字段名称': '谈话类别',
                    '字段类型': '下拉选择',
                    '必填': '是',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$TalkType']['options'] if opt['text']]),
                    '填写说明': '选择谈话类型',
                    '系统字段': 'TalkHeartPage$TalkType'
                },
                {
                    '字段名称': '谈话类别补充',
                    '字段类型': '文本',
                    '必填': '否',
                    '可选值': '',
                    '填写说明': '选择"其他"时填写具体类型',
                    '系统字段': 'TalkHeartPage$TalkTypeInput'
                },
                {
                    '字段名称': '谈话类型',
                    '字段类型': '下拉选择',
                    '必填': '是',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$TalkType1']['options'] if opt['text']]),
                    '填写说明': '固定选择"一对一谈话"',
                    '系统字段': 'TalkHeartPage$TalkType1'
                },
                {
                    '字段名称': '困惑和问题',
                    '字段类型': '下拉选择',
                    '必填': '是',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$QuesRemark']['options'] if opt['text']]),
                    '填写说明': '选择具体问题类型',
                    '系统字段': 'TalkHeartPage$QuesRemark'
                }
            ])
            
            # 谈话主题（必填）
            template_data.append({
                '字段名称': '谈话主题',
                '字段类型': '多选',
                '必填': '是',
                '可选值': '\n'.join([
                    self.form_data['fields'][f'TalkHeartPage$ConversationTopic${i}']['label']
                    for i in range(7)
                    if self.form_data['fields'][f'TalkHeartPage$ConversationTopic${i}']['label']
                ]),
                '填写说明': '至少选择一项，多个用逗号分隔',
                '系统字段': 'ConversationTopic'
            })
            
            # 谈话内容（必填）
            template_data.extend([
                {
                    '字段名称': '面谈地点及方式',
                    '字段类型': '下拉选择',
                    '必填': '是',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$TalkAddressValue']['options'] if opt['text']]),
                    '填写说明': '选择谈话地点',
                    '系统字段': 'TalkHeartPage$TalkAddressValue'
                },
                {
                    '字段名称': '谈心概况',
                    '字段类型': '文本区域',
                    '必填': '是',
                    '可选值': '',
                    '填写说明': '详细记录谈话内容',
                    '系统字段': 'TalkHeartPage$TalkContenJS'
                }
            ])
            
            # 帮扶措施（必填）
            template_data.extend([
                {
                    '字段名称': '提供帮助1',
                    '字段类型': '下拉选择',
                    '必填': '是',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$TGHelp1']['options'] if opt['text']]),
                    '填写说明': '选择帮扶措施类型',
                    '系统字段': 'TalkHeartPage$TGHelp1'
                }
            ])
            
            # 谈话人身份（必填）
            template_data.append({
                '字段名称': '与谈话人关系',
                '字段类型': '下拉选择',
                '必填': '是',
                '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$Relation']['options'] if opt['text']]),
                '填写说明': '选择谈话人身份',
                '系统字段': 'TalkHeartPage$Relation'
            })
            
            # 谈话结果
            template_data.append({
                '字段名称': '谈话结果',
                '字段类型': '多选',
                '必填': '否',
                '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$TalkResult']['options'] if opt['text']]),
                '填写说明': '可多选，用逗号分隔',
                '系统字段': 'TalkHeartPage$TalkResult'
            })
            
            # 是否关注
            template_data.extend([
                {
                    '字段名称': '是否重点关注',
                    '字段类型': '单选',
                    '必填': '否',
                    '可选值': 'IsCareY=是,IsCareN=否',
                    '填写说明': '选择是否需要关注',
                    '系统字段': 'TalkHeartPage$IsCare'
                },
                {
                    '字段名称': '关注等级',
                    '字段类型': '下拉选择',
                    '必填': '否',
                    '可选值': '\n'.join([opt['text'] for opt in self.form_data['fields']['TalkHeartPage$CareType']['options'] if opt['text']]),
                    '填写说明': '如需关注，选择关注等级',
                    '系统字段': 'TalkHeartPage$CareType'
                },
                {
                    '字段名称': '关注开始时间',
                    '字段类型': '文本',
                    '必填': '否',
                    '可选值': '',
                    '填写说明': '填写数字（月份）',
                    '系统字段': 'TalkHeartPage$CareBeg'
                },
                {
                    '字段名称': '关注结束时间',
                    '字段类型': '文本',
                    '必填': '否',
                    '可选值': '',
                    '填写说明': '填写数字（月份）',
                    '系统字段': 'TalkHeartPage$CareEnd'
                }
            ])
            
            # 其他备注
            template_data.append({
                '字段名称': '备注',
                '字段类型': '文本区域',
                '必填': '否',
                '可选值': '',
                '填写说明': '其他需要说明的情况',
                '系统字段': 'TalkHeartPage$Memo'
            })
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '谈心记录'
            
            # 写入基本列
            headers = ['字段名称', '字段类型', '必填', '可选值', '填写说明']
            # 添加填写内容列
            for i in range(1, record_count + 1):
                headers.append(f'记录{i}')
            # 添加系统字段列
            headers.append('系统字段')
            
            ws.append(headers)
            
            # 写入字段信息
            for row_data in template_data:
                # 准备基本列数据
                row = [
                    row_data['字段名称'],
                    row_data['字段类型'],
                    row_data['必填'],
                    row_data['可选值'],
                    row_data['填写说明']
                ]
                
                # 如果是姓名字段且有学生名单，自动填入学生姓名
                if row_data['字段名称'] == '姓名' and student_list:
                    for i in range(record_count):
                        if i < len(student_list):
                            # 从学生信息中提取姓名
                            name = None
                            for key in student_list[i]:
                                if '姓名' in key:
                                    name = student_list[i][key]
                                    break
                            row.append(name or '')
                        else:
                            row.append('')
                else:
                    # 添加空的填写内容列
                    for _ in range(record_count):
                        row.append('')
                
                # 添加系统字段
                row.append(row_data['系统字段'])
                ws.append(row)
            
            # 设置样式
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 应用样式
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 字段名称
            ws.column_dimensions['B'].width = 10  # 字段类型
            ws.column_dimensions['C'].width = 8   # 必填
            ws.column_dimensions['D'].width = 30  # 可选值
            ws.column_dimensions['E'].width = 20  # 填写说明
            
            # 设置填写内容列的宽度
            for i in range(record_count):
                col = get_column_letter(6 + i)  # 从第6列开始
                ws.column_dimensions[col].width = 30
            
            # 设置系统字段列
            sys_col = get_column_letter(6 + record_count)
            ws.column_dimensions[sys_col].width = 30
            ws.column_dimensions[sys_col].hidden = True
            
            # 冻结首行和前五列
            ws.freeze_panes = 'F2'
            
            # 保存文件
            wb.save(output_file)
            return True
            
        except Exception as e:
            print(f'生成模板失败: {str(e)}')
            return False

    def run(self):
        """运行自动填写流程"""
        try:
            # 检查是否已登录
            if not self.is_logged_in and not self.login():
                print('❌ 登录失败')
                return False
            
            # 导航到表单页面
            if not self._navigate_to_form():
                return False
            
            # 选择要使用的模板
            print('\n请选择要使用的模板文件：')
            template_files = []
            
            # 查找所有模板文件
            for file in glob.glob('谈心记录模板_*.xlsx'):
                template_files.append(file)
                print(f'{len(template_files)}. {file}')
            
            if not template_files:
                print('❌ 未找到任何模板文件')
                print('请先运行数据提取工具生成模板')
                return False
            
            # 让用户选择模板
            while True:
                choice = input('\n请选择要使用的模板文件 (输入序号): ').strip()
                try:
                    idx = int(choice) - 1
                    if 0 <= idx < len(template_files):
                        self.excel_file = template_files[idx]
                        print(f'✓ 已选择: {self.excel_file}')
                        break
                    else:
                        print('无效的选择，请重新输入')
                except ValueError:
                    print('请输入有效的数字')
            
            # 加载Excel数据
            if not self._load_excel_data():
                return False
            
            # 开始提交记录
            total = len(self.data)
            success = 0
            failed_records = []
            
            print(f'\n开始处理 {total} 条记录...')
            for i, record in enumerate(self.data, 1):
                student_name = record.get('TalkHeartPage$StuName', '未知学生')
                print(f'\n处理第 {i}/{total} 条记录: {student_name}')
                
                try:
                    if self.submit_record(record):
                        success += 1
                        print(f'等待3秒后处理下一条记录...')
                        time.sleep(3)
                    else:
                        failed_records.append({
                            'name': student_name,
                            'reason': '提交失败，可能是表单验证未通过'
                        })
                except Exception as e:
                    failed_records.append({
                        'name': student_name,
                        'reason': str(e)
                    })
            
            # 打印结果
            print(f'\n处理完成: 成功 {success}/{total} 条记录')
            if failed_records:
                print('\n失败记录:')
                for record in failed_records:
                    print(f'- {record["name"]}: {record["reason"]}')
            return success > 0
            
        except Exception as e:
            print(f'执行出错: {str(e)}')
            return False
        finally:
            input('按回车键退出...')
            if self.driver:
                self.driver.quit()

    def submit_record(self, record):
        """提交单条记录"""
        try:
            print('- 等待表单加载...')
            # 等待表单加载完成
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, 'form1'))
            )
            time.sleep(2)

            # 首先设置谈话类型为"一对一谈话"
            print('- 设置谈话类型...')
            try:
                talk_type = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.NAME, 'TalkHeartPage$TalkType1'))
                )
                Select(talk_type).select_by_value('08031001')  # 一对一谈话的值
                time.sleep(2)
            except Exception as e:
                print(f'设置谈话类型时出错: {str(e)}')
                return False

            # 填写基本信息
            print('- 开始填写表单...')
            for field, value in record.items():
                try:
                    print(f'  正在填写 {field}: {value}')
                    if field == 'TalkHeartPage$StuName':
                        if not self._handle_student_name(value):
                            return False
                        time.sleep(1)
                    elif field == 'TalkHeartPage$TalkType1':
                        continue  # 已经设置过了
                    elif field == 'ConversationTopic':
                        if not self._handle_conversation_topics(value):
                            return False
                        time.sleep(1)
                    elif field == 'TalkHeartPage$IsCare':
                        self._handle_radio(field, value)
                        time.sleep(1)
                    else:
                        self._fill_field(field, value)
                        time.sleep(1)
                except Exception as e:
                    print(f'填写字段 {field} 时出错: {str(e)}')
                    return False

            print('✓ 表单填写完成')
            
            # 点击保存按钮
            try:
                print('- 点击保存按钮...')
                save_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.ID, 'Save'))
                )
                save_btn.click()
                time.sleep(2)
                
                # 处理弹窗
                alert = self.driver.switch_to.alert
                alert_text = alert.text
                alert.accept()
                
                if '成功' in alert_text:
                    print('✓ 提交成功')
                    return True
                else:
                    print(f'❌ 提交失败: {alert_text}')
                    return False
                
            except Exception as e:
                print(f'保存表单时出错: {str(e)}')
                return False
                
        except Exception as e:
            print(f'表单处理出错: {str(e)}')
            return False

    def close_browser(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()
            self.driver = None

    def check_environment(self):
        """检查并创建必要的运行环境"""
        try:
            # 检查配置文件目录
            if not os.path.exists('config'):
                os.makedirs('config')
            
            # 检查模板目录
            if not os.path.exists('templates'):
                os.makedirs('templates')
            
            # 检查日志目录
            if not os.path.exists('logs'):
                os.makedirs('logs')
                
        except Exception as e:
            print(f'创建目录失败: {str(e)}')

def _extract_fields(driver, form, target_dict, fields=None):
    """提取字段信息的辅助函数"""
    if fields is None:
        fields = form.find_elements(By.CSS_SELECTOR, 
            'input:not([type="hidden"]), select, textarea')
    
    for field in fields:
        try:
            name = field.get_attribute('name')
            if name and name not in target_dict:
                # 重新获取元素，避免stale element
                field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.NAME, name))
                )
                
                label = _get_field_label(driver, field)
                
                field_info = {
                    'type': field.get_attribute('type') or field.tag_name,
                    'value': field.get_attribute('value') or '',
                    'label': label,
                    'required': field.get_attribute('required') == 'true',
                    'validation': {
                        'max_length': field.get_attribute('maxlength'),
                        'pattern': field.get_attribute('pattern'),
                        'min': field.get_attribute('min'),
                        'max': field.get_attribute('max')
                    }
                }
                
                # 处理select元素
                if field.tag_name == 'select':
                    options = field.find_elements(By.TAG_NAME, 'option')
                    field_info['options'] = [
                        {
                            'value': opt.get_attribute('value'),
                            'text': opt.text.strip(),
                            'selected': opt.is_selected()
                        }
                        for opt in options
                    ]
                    field_info['type'] = ('select-multiple' 
                        if field.get_attribute('multiple') else 'select-one')
                
                # 处理radio和checkbox
                elif field.get_attribute('type') in ['radio', 'checkbox']:
                    field_info['checked'] = field.is_selected()
                    if field.get_attribute('type') == 'radio':
                        other_options = form.find_elements(
                            By.CSS_SELECTOR, 
                            f'input[type="radio"][name="{name}"]'
                        )
                        field_info['options'] = [
                            {
                                'value': opt.get_attribute('value'),
                                'text': _get_field_label(driver, opt),
                                'checked': opt.is_selected()
                            }
                            for opt in other_options
                        ]
                
                field_info['visible'] = field.is_displayed()
                field_info['enabled'] = field.is_enabled()
                
                placeholder = field.get_attribute('placeholder')
                if placeholder:
                    field_info['placeholder'] = placeholder
                
                target_dict[name] = field_info
                print(f'✓ 已提取: {name} ({label})')
        except Exception as e:
            print(f'提取字段 {name} 时出错: {str(e)}')
            continue

def _get_field_label(driver, field):
    """获取表单字段的标签文本"""
    label_text = ''
    try:
        # 方法1：通过 for 属性查找
        field_id = field.get_attribute('id')
        if field_id:
            label = driver.find_element(By.CSS_SELECTOR, f'label[for="{field_id}"]')
            if label:
                label_text = label.text.strip()
        
        # 方法2：查找父级 label
        if not label_text:
            parent_label = field.find_element(By.XPATH, './ancestor::label')
            if parent_label:
                label_text = parent_label.text.strip()
        
        # 方法3：查找前面的 label
        if not label_text:
            prev_label = field.find_element(
                By.XPATH, 
                './preceding::label[1]'
            )
            if prev_label:
                label_text = prev_label.text.strip()
        
        # 方法4：使用 aria-label 属性
        if not label_text:
            label_text = field.get_attribute('aria-label') or ''
            
        # 方法5：使用 title 属性
        if not label_text:
            label_text = field.get_attribute('title') or ''
            
        return label_text.strip()
        
    except:
        return ''

def show_menu():
    """显示菜单"""
    extractor = FormExtractor()
    
    try:
        print('\n' + '=' * 50)
        print('表单字段提取工具')
        print('=' * 50)
        
        # 显示已加载的数据状态
        if extractor.form_data:
            print('✓ 已加载表单字段信息')
        if extractor.student_list:
            print(f'✓ 已加载 {len(extractor.student_list)} 条学生记录')
        
        print('正在准备登录...')
        
        if not extractor.login():
            print('❌ 登录失败，程序退出')
            return
        
        while True:
            print('\n' + '=' * 50)
            print('功能选项：')
            print('1. 提取表单字段信息')
            print('2. 提取学生名单')
            print('3. 生成谈心记录模板')
            print('4. 自动填写谈心记录')
            print('5. 退出程序')
            print('=' * 50)
            
            choice = input('请选择功能 (1-5): ').strip()
            
            if choice == '1':
                if extractor.extract_form_fields():
                    print('✓ 表单字段提取完成')
                else:
                    print('❌ 表单字段提取失败')
            elif choice == '2':
                if extractor.extract_student_list():
                    print('✓ 学生名单提取完成')
                else:
                    print('❌ 学生名单提取失败')
            elif choice == '3':
                if extractor.generate_template_by_care_level():
                    print('✓ 模板生成完成')
                else:
                    print('❌ 模板生成失败')
            elif choice == '4':
                if extractor.run():
                    print('✓ 自动填写完成')
                else:
                    print('❌ 自动填写失败')
            elif choice == '5':
                print('\n感谢使用！')
                break
            else:
                print('\n无效的选择，请重新输入')
            
            print('\n是否继续使用其他功能？(y/n)')
            if input().lower() != 'y':
                print('\n感谢使用！')
                break
        
    except Exception as e:
        print(f'\n程序发生错误: {str(e)}')
    finally:
        print('\n是否关闭浏览器？(y/n)')
        if input().lower() == 'y':
            extractor.close_browser()
        input('\n按回车键退出...')

if __name__ == '__main__':
    show_menu() 